import io
import csv
import fitz  # pymupdf — для PDF
import openpyxl  # для Excel
from fastapi import APIRouter, UploadFile, File, Depends
from app.database import supabase
from app.auth import get_current_user
from app.whatsapp import (
    start_session, stop_session, get_session_status,
    get_qr_code, setup_webhook as waha_setup_webhook
)
from app.config import BACKEND_URL

router = APIRouter()


@router.get("/api/bots")
def get_bots(user: dict = Depends(get_current_user)):
    """Получает список ботов текущего пользователя."""
    result = supabase.table("bots").select("*").eq("user_id", user["id"]).execute()
    return {"bots": result.data}


@router.post("/api/bots")
def create_bot(data: dict, user: dict = Depends(get_current_user)):
    """Создаёт нового бота для текущего пользователя."""
    result = supabase.table("bots").insert({
        "name": data.get("name", "Новый бот"),
        "system_prompt": data.get("system_prompt", ""),
        "whatsapp_session": data.get("whatsapp_session", "default"),
        "user_id": user["id"],
    }).execute()

    return {"bot": result.data[0]}


@router.put("/api/bots/{bot_id}")
def update_bot(bot_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Обновляет настройки бота (промпт, имя и т.д.)."""
    # Безопасно: не разрешаем менять user_id и id
    safe_data = {k: v for k, v in data.items() if k not in ("id", "user_id")}

    result = supabase.table("bots") \
        .update(safe_data) \
        .eq("id", bot_id) \
        .eq("user_id", user["id"]) \
        .execute()

    if not result.data:
        return {"error": "Бот не найден"}

    return {"bot": result.data[0]}


def extract_text_from_pdf(content: bytes) -> str:
    """Извлекает текст из PDF файла."""
    doc = fitz.open(stream=content, filetype="pdf")
    text_parts = []
    for page in doc:
        text_parts.append(page.get_text())
    doc.close()
    return "\n".join(text_parts)


def extract_rows_from_xlsx(content: bytes) -> list:
    """
    Извлекает строки из Excel файла.
    Каждая строка данных возвращается как отдельный элемент с заголовками столбцов.
    Формат: "Столбец1: Значение1 | Столбец2: Значение2 | ..."
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    rows_with_headers = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = []
        first_row = True

        for row in ws.iter_rows(values_only=True):
            cells = [str(cell).strip() if cell is not None else "" for cell in row]

            # Пропускаем полностью пустые строки
            if not any(cells):
                continue

            if first_row:
                # Первая непустая строка — заголовки
                headers = cells
                first_row = False
                continue

            # Формируем строку с заголовками: "Название: iPhone | Цена: 100000 | ..."
            parts = []
            for i, cell in enumerate(cells):
                if cell:
                    header = headers[i] if i < len(headers) and headers[i] else f"Столбец {i+1}"
                    parts.append(f"{header}: {cell}")

            if parts:
                rows_with_headers.append(" | ".join(parts))

    wb.close()
    return rows_with_headers


def extract_rows_from_csv(content: bytes) -> list:
    """
    Извлекает строки из CSV файла.
    Каждая строка данных возвращается с заголовками столбцов.
    """
    text = content.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows_with_headers = []
    headers = []
    first_row = True

    for row in reader:
        cells = [cell.strip() for cell in row]

        if not any(cells):
            continue

        if first_row:
            headers = cells
            first_row = False
            continue

        parts = []
        for i, cell in enumerate(cells):
            if cell:
                header = headers[i] if i < len(headers) and headers[i] else f"Столбец {i+1}"
                parts.append(f"{header}: {cell}")

        if parts:
            rows_with_headers.append(" | ".join(parts))

    return rows_with_headers


@router.post("/api/bots/{bot_id}/knowledge")
async def upload_knowledge(bot_id: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """
    Загружает файл с базой знаний.
    Поддерживаемые форматы: TXT, PDF, CSV, XLSX.
    Разбивает на чанки и сохраняет в таблицу knowledge_chunks.
    """
    content = await file.read()
    filename = file.filename.lower() if file.filename else ""

    chunks = []

    # Для таблиц — каждая строка = отдельный чанк с заголовками
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        rows = extract_rows_from_xlsx(content)
        for row in rows:
            chunks.append({"bot_id": bot_id, "content": row})

    elif filename.endswith(".csv"):
        rows = extract_rows_from_csv(content)
        for row in rows:
            chunks.append({"bot_id": bot_id, "content": row})

    else:
        # PDF и TXT — разбиваем текст на чанки по ~500 символов
        if filename.endswith(".pdf"):
            text = extract_text_from_pdf(content)
        else:
            text = content.decode("utf-8", errors="ignore")

        chunk_size = 500
        for i in range(0, len(text), chunk_size):
            chunk = text[i:i + chunk_size].strip()
            if chunk:
                chunks.append({"bot_id": bot_id, "content": chunk})

    if chunks:
        # Удаляем старые чанки этого бота
        supabase.table("knowledge_chunks") \
            .delete() \
            .eq("bot_id", bot_id) \
            .execute()

        # Вставляем новые
        supabase.table("knowledge_chunks").insert(chunks).execute()

    return {"chunks_count": len(chunks), "format": filename.split(".")[-1] if "." in filename else "txt"}


@router.post("/api/bots/{bot_id}/images")
async def upload_image(bot_id: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """
    Загружает картинку в Supabase Storage для бота.
    Бот сможет отправлять её клиентам.
    """
    content = await file.read()
    filename = file.filename or "image.jpg"

    # Путь в Supabase Storage: bot_images/{bot_id}/{filename}
    storage_path = f"{bot_id}/{filename}"

    # Загружаем в Supabase Storage (бакет bot_images)
    supabase.storage.from_("bot_images").upload(storage_path, content)

    # Получаем публичный URL
    public_url = supabase.storage.from_("bot_images").get_public_url(storage_path)

    # Сохраняем URL в таблицу knowledge_chunks с пометкой [IMAGE]
    supabase.table("knowledge_chunks").insert({
        "bot_id": bot_id,
        "content": f"[IMAGE:{public_url}] {filename}",
    }).execute()

    return {"url": public_url, "filename": filename}


@router.get("/api/bots/{bot_id}/images")
def get_images(bot_id: str, user: dict = Depends(get_current_user)):
    """Получает список загруженных картинок бота."""
    result = supabase.table("knowledge_chunks") \
        .select("id, content") \
        .eq("bot_id", bot_id) \
        .like("content", "[IMAGE:%") \
        .execute()

    images = []
    for item in result.data:
        # Парсим URL из формата [IMAGE:url] filename
        content = item["content"]
        url = content.split("]")[0].replace("[IMAGE:", "")
        name = content.split("] ")[-1] if "] " in content else "image"
        images.append({"id": item["id"], "url": url, "name": name})

    return {"images": images}


# --- WhatsApp (WAHA) ---

@router.post("/api/bots/{bot_id}/whatsapp/start")
async def whatsapp_start(bot_id: str, user: dict = Depends(get_current_user)):
    """Запускает сессию WhatsApp для бота и настраивает webhook."""
    # Используем bot_id как имя сессии
    session_name = f"bot_{bot_id}"

    # Запускаем сессию
    result = await start_session(session_name)

    # Настраиваем webhook для приёма сообщений
    webhook_url = f"{BACKEND_URL}/webhook/whatsapp"
    await waha_setup_webhook(session_name, webhook_url)

    # Сохраняем имя сессии в бота
    supabase.table("bots") \
        .update({"whatsapp_session": session_name}) \
        .eq("id", bot_id) \
        .eq("user_id", user["id"]) \
        .execute()

    return {"status": "started", "session": session_name, "result": result}


@router.post("/api/bots/{bot_id}/whatsapp/stop")
async def whatsapp_stop(bot_id: str, user: dict = Depends(get_current_user)):
    """Останавливает сессию WhatsApp для бота."""
    session_name = f"bot_{bot_id}"
    result = await stop_session(session_name)

    supabase.table("bots") \
        .update({"whatsapp_session": ""}) \
        .eq("id", bot_id) \
        .eq("user_id", user["id"]) \
        .execute()

    return {"status": "stopped", "result": result}


@router.get("/api/bots/{bot_id}/whatsapp/status")
async def whatsapp_status(bot_id: str, user: dict = Depends(get_current_user)):
    """Получает статус подключения WhatsApp."""
    session_name = f"bot_{bot_id}"
    status = await get_session_status(session_name)
    return status


@router.get("/api/bots/{bot_id}/whatsapp/qr")
async def whatsapp_qr(bot_id: str, user: dict = Depends(get_current_user)):
    """Получает QR-код для подключения WhatsApp."""
    session_name = f"bot_{bot_id}"
    qr_data = await get_qr_code(session_name)
    return qr_data
